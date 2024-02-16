

'''
##########################################################

    Author:
           James R Drummond 2024

    License:
    
    Usage:
    
    This is a simple python program to produce the BOM and layout
    suitable from extract FUSION 360 files to send to jlcpcb.com
    
    python extract [zip] [-t through-hole] [-b previous bom] [-B]
    
    zip:
        The gerber zip from FUSION 360 (default "archive.zip"
        "your PCB layout" > MANUFACTURING > MANUFACTURING > Export gerger....
        
    -t through-hole:
        If you only need the SMT, your don't need this file
        To add the through-hole parts run
        "your PCB layout" > AUTOMATION > AUTOMATION > Run ULP
          then run "mount_smd_tht.ulp" (use the defaults) which
          produces the "tht" file needed
          
    -b previous bom
        tries allocated the jclpcb part # from the original to the current BOM
        
    -B
        gets the existing bom and verifirs that it has not changed
        
    
    
    The result is the BOM and layout needed.
    
    At the end the BOM must be populated with part numbers, 
                      usually from jlcpcb.com/parts.
    
    The layout might need to be tweaked because the jclpcb system doesn't
    always get the orientation correct.  The best thing is to upload to 
    jclpcb and look at their version on the layout picture and then adjust the
    layout file.
    
          
'''

import os
import zipfile
import csv
import openpyxl 
import sys
import io
from argparse import ArgumentParser
from time import strftime

if __name__ == "__main__":

    def abc():
        return "archive"
##########################################################
#
#          Startup
#
##########################################################
    prog  =  "usage %s [gerber zip] [-t through-hole csv] [-b original bom]" % sys.argv[0]
    header = "program to extract file from FUSION 360 to jclpcb.com"
    parser = ArgumentParser(prog=prog,description=header)
    parser.add_argument("archive",nargs='?',default="archive",help='The gerber zip file')
    parser.add_argument("-t", "--through-hole",action="store",dest="tht",default=None ,help="through hole csv file")
    parser.add_argument("-b", "--bom", dest="original_bom", action='store',default=None,help="original bom for updating")    
    parser.add_argument("-B", "--BOM", dest="verify_bom", action='store_true',default=None,help="verify the BOM is unchanged")    
    args = parser.parse_args()
    archive      = args.archive
    tht          = args.tht
    original_bom = args.original_bom
    verify_bom   = args.verify_bom
    
    print (prog, strftime("%a %b %d, %H:%M"))
       
    a = False 
    try:
        a = archive[:-4].lower() == '.zip'
    except:
        pass
        
    if not a:  archive += '.zip'
        
    try:
        f = open(archive,'r')    
        f.close()
    except:
        print("Cannot open ",archive)
        sys.exit()

##########################################################
#
#         extract from zip
#
##########################################################
    
    files = []
    with zipfile.ZipFile(archive) as z:
        for filename in z.namelist():
            a = filename.split('/')
            if a[-2] != 'Assembly': continue
            s = None
            try:
                with z.open(filename,) as f: s = f.read()
            except:
                pass
            if not (s is None): files.append({'name':a[-1]      , 'data':s.decode('utf-8')})
            
##########################################################
#
#          check whether files are reasonable
#
##########################################################
    bom = []
    top = []
    bottom = []
    print("Input files: ")
    for file in files:
        print('    ',file['name'])
        if 'back' in file['name']:
            bottom.append(file)
        elif 'front' in file['name']: 
            top.append(file)
        else:
            bom.append(file)
    if len(bom) != 1:
        print("One bom is not unique (none or multiple")
        for file in files:  print(file['name'])
        sys.exit()
    
    if len(top) == 1:
        top = top[0]
    else:
        print("Layouts Problem 0a")
        for file in files:  print(file['name'])
        sys.exit()

    if len(bottom) == 1:
        bottom = bottom[0]
    else:
        print("Layouts Problem 0b")
        for file in files:  print(file['name'])
        sys.exit()

##########################################################
#
#          found a bom and possibly one top, bottom layout
#          find which is top or bottom
#
##########################################################
    
    if (len(top['data']) != 0) and (len(bottom['data']) == 0):
        layout = top
        side = 'TOP'
    elif (len(top['data']) == 0) and (len(botom['data']) != 0):
        layout = bottom
        side = 'BOTTOM'
    else:
        print("Layouts Problem 1" )
        for file in files:  print(file['name'])
        sys.exit()
##########################################################
#
#          look for through-hole parts
#          and massage them to look it like the gerber layout
#
##########################################################
   
    

    
    if not tht is None:
        print(tht, "is the file of the through-hole data")
        if tht[-4].lower() != ".csv":  tht += '.csv'
        s = None
        try:
            with open(tht,'r') as f: s = f.read()
        except:
            print("Unable to find through-hole file ",tht)
            sys.exit()
        
        for ss in s.splitlines():
            a = ss.split(',')
            b = [a[0]] + a[3:6] + [a[0]] + [a[2]]
            c = " ".join(b)
            layout['data'] += c + '\n'
            print("Added the through-hole: ",c )

     
       
##########################################################
#
#          rewrite the layout
#
##########################################################
      
    
    r = layout['data'].splitlines()    
    devices = []
    name = layout['name'][:-4]  + ' layout.xlsx'
    print("Creating ",name)
    workbook =   openpyxl.Workbook()
    page = workbook.active
    page['A1'] = 'Designator'
    page['B1'] = 'Mid X'
    page['C1'] = 'Mid Y'
    page['D1'] = 'Layer'
    page['E1'] = 'Rotation'
    i = 2
    for row in r:

        a = row.split()
        try:
            devices.append(a[0])
        except:
            continue
        page['A' + str(i)] = a[0]
        page['B' + str(i)] = a[1]
        page['C' + str(i)] = a[2]
        page['D' + str(i)] = side
        page['E' + str(i)] = a[3]
        i += 1
    workbook.save(name)

##########################################################
#
#          find the bom columns 
#
##########################################################
    bom = bom[0]
    r = bom['data'].splitlines()
    i = 0
    comment = [r[2].index('Value'),r[2].index('Device')]
    designator = [r[2].index('Parts'),r[2].index('Description')]
    footprint = [r[2].index('Package'),r[2].index('Parts')]
    if "jclpcb" in r[2].lower():
        z = r[2].lower().index["jclpcb"]
        i = z
        while (r[2][i] != " ") and (i != 0): i -= 1
        a = min(0,i+1)
        i  = z
        while (r[2][i] != " ") and (i != len(r[2])): i += 1
        partNo = [a,i]
    else:
        partNo = None
    name = bom['name'][:-4] + ' bom.xlsx'


    if verify_bom: 
##########################################################
#
#          verify the bom (and the part #s are let alone)
#
##########################################################
        
            
        print("Verifying ",name)
        try:
            workbook =   openpyxl.load_workbook(filename=name)
        except:
            raise NameError("Unable to find old bom: " + name)
        page =  workbook.active
        i = 2
        for row in r[3:]:
            c = row[comment[0]:comment[1]]
            d = row[designator[0]:designator[1]]
            f = row[footprint[0]:footprint[1]]
            d1 = d.replace(',',' ').split()
            found = False
            for item in d1:
                found = found or (item.strip() in devices)
            if found:
                A = page['A' + str(i)].value if not page['A' + str(i)].value is None else  ""
                B = page['B' + str(i)].value if not page['B' + str(i)].value is None else  ""
                C = page['C' + str(i)].value if not page['C' + str(i)].value is None else  ""
                if  A != c.strip() or \
                    B != d.strip() or \
                    C != f.strip() :  
                        raise NameError("Item %d is different" % (i-1))
                i += 1
        print("Verified ",name)
    else:
##########################################################
#
#          find the previous bom
#
##########################################################
    
        parts = {}
        if not original_bom is None:
            if not original_bom[:-5] == ".xlsx": original_bom += ".xlsx"
            print("The original bom is ",original_bom)
            workbook =   openpyxl.load_workbook(original_bom)
            page = workbook.active
            i = 2
            while not page['B' + str(i)].value is None:
                A = page['A' + str(i)].value if not page['A' + str(i)].value is None else  ""
                B = page['B' + str(i)].value if not page['B' + str(i)].value is None else  ""
                C = page['C' + str(i)].value if not page['C' + str(i)].value is None else  ""
                D = page['D' + str(i)].value if not page['C' + str(i)].value is None else  ""
                try:
                    parts[A+B+C]  = str(D).strip()
                except:
                    break
                i += 1
            print("%d items in original bom" % (i-2))
 
     
##########################################################
#
#          rewrite the bom
#
##########################################################
    
        
            
        print("Creating ",name)
        workbook =   openpyxl.Workbook()
        page = workbook.active
        page['A1'] = 'Comment'
        page['B1'] = 'Designator'
        page['C1'] = 'Footprint'
        page['D1'] = 'JCLPCB Part #'
        i = 2
        for row in r[3:]:
            c = row[comment[0]:comment[1]]
            d = row[designator[0]:designator[1]]
            e = row[footprint[0]:footprint[1]]
            d1 = d.replace(',',' ').split()
            f = ""
            if not partNo is None:  f = row[partNo[0]:partNo[1]]
            found = False
            for item in d1:
                found = found or (item.strip() in devices)
            if found:
                page['A' + str(i)] = c.strip()
                page['B' + str(i)] = d.strip()
                page['C' + str(i)] = e.strip()
                g = ""
                g = parts[page['A' + str(i)].value + \
                    page['B' + str(i)].value + \
                        page['C' + str(i)].value]
                try:
                    g = parts[page['A' + str(i)].value + \
                        page['B' + str(i)].value + \
                        page['C' + str(i)].value]
                except:
                    pass
                if (f != g) and (f != "") and (g != ""):
                    print("Conflict on part #, part %s parts %s and %s - using %s" \
                            %(d,f,g,g) )
                elif (f != "") and (g == ""):
                    g = f
                page['D' + str(i)] = g.strip()
                i += 1
        workbook.save(name)
 